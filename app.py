import streamlit as st
import json
import fitz
import pandas as pd
import re
from io import BytesIO

st.set_page_config(page_title="Comparador PDF vs JSON", layout="wide")

st.markdown("""
<style>
.semaforo-verde    { background:#1a472a; color:#a9f5c0; padding:18px 24px; border-radius:12px; font-size:22px; font-weight:bold; text-align:center; margin-bottom:10px; }
.semaforo-amarillo { background:#7a6000; color:#ffe680; padding:18px 24px; border-radius:12px; font-size:22px; font-weight:bold; text-align:center; margin-bottom:10px; }
.semaforo-rojo     { background:#6b1a1a; color:#ffaaaa; padding:18px 24px; border-radius:12px; font-size:22px; font-weight:bold; text-align:center; margin-bottom:10px; }
.sin-pareja        { background:#3a2000; color:#ffcc80; padding:10px 16px; border-radius:8px; margin-bottom:6px; font-size:14px; }
</style>
""", unsafe_allow_html=True)


# =========================================================
# FUNCIONES AUXILIARES
# =========================================================

def limpiar_texto(txt):
    if txt is None:
        return ""
    return str(txt).strip()

def limpiar_upper(txt):
    return limpiar_texto(txt).upper()

def convertir_a_float(valor):
    try:
        valor = str(valor).replace(",", ".").strip()
        return round(float(valor), 2)
    except:
        return None

def convertir_a_int(valor):
    try:
        return int(round(float(str(valor).replace(",", ".").strip())))
    except:
        return None

def a_euro(valor):
    num = convertir_a_float(valor)
    if num is None:
        return ""
    return f"{num:.2f} €"


def convertir_a_num(valor):
    """Convierte un valor numérico devolviendo int si es entero, o float si tiene decimales.
    Útil para tamaños del PDF/Excel que pueden ser '400' o '35.2'."""
    try:
        f = float(str(valor).replace(",", ".").strip())
        # Si es entero exacto, devolver int para mostrar '400' en vez de '400.0'
        if f == int(f):
            return int(f)
        # Si tiene decimales reales, redondear a 2 para evitar floats sucios
        return round(f, 2)
    except:
        return None

def crear_excel_en_memoria(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Diferencias")
    output.seek(0)
    return output.getvalue()

def son_textos_distintos(a, b):
    return limpiar_upper(a) != limpiar_upper(b)

def son_numeros_distintos(a, b):
    fa = convertir_a_float(a)
    fb = convertir_a_float(b)
    if fa is None and fb is None:
        return False
    if fa is None or fb is None:
        return True
    return fa != fb


UMBRAL_PRECIO_LINEA = 4000.0


# =========================================================
# PARSEAR JSON
# =========================================================

def parsear_json(data):
    # Campos de cabecera:
    #  - "importe" del JSON = suma de muebles (aparece también en el PDF como IMPORTE)
    #  - "ivaCabinetsPorcentaje" = 21% del importe de muebles (aparece en PDF como "I.V.A. 21% en muebles")
    #  - El Total que aparece en el PDF = IMPORTE + I.V.A. de muebles
    # Los campos "iva" y "total" del JSON suman IVAs de otras categorías (encimeras, electrodomésticos,
    # equipamientos) que NO salen en este PDF, por eso no los usamos aquí.
    importe_json = convertir_a_float(data.get("importe", 0))
    iva_json     = convertir_a_float(data.get("ivaCabinetsPorcentaje"))
    if iva_json is None:
        iva_json = convertir_a_float(data.get("iva", 0))  # fallback

    if importe_json is not None and iva_json is not None:
        total_json = round(importe_json + iva_json, 2)
    else:
        total_json = convertir_a_float(data.get("total", 0))  # fallback

    resumen = {
        "pedido":   limpiar_texto(data.get("orderCode", "")),
        "cliente":  limpiar_texto(data.get("customerName", "")),
        "tienda":   limpiar_texto(data.get("storeName", "")),
        "proyecto": limpiar_texto(data.get("projectName", "")),
        "importe":  importe_json,
        "iva":      iva_json,
        "total":    total_json,
    }
    lineas = []
    for item in data.get("cabinets", []):
        size = item.get("size") or {}
        size_x = convertir_a_num(size.get("x"))
        size_y = convertir_a_num(size.get("y"))
        size_z = convertir_a_num(size.get("z"))

        price_total = convertir_a_float(item.get("priceTotal"))
        if price_total is None:
            price_total = convertir_a_float(item.get("total"))

        # Modelo y material de la puerta/frente
        doors = item.get("doors") or {}
        model_door    = limpiar_texto(item.get("modelDoor") or doors.get("name") or "")
        material_door = limpiar_texto(item.get("materialDoor") or doors.get("material") or "")

        # Material del armazón (muebles, regletas con armazón)
        material_cabinet = limpiar_texto(item.get("materialCabinet") or "")

        # Material genérico (complementos, costados, zócalos → "Acabado" en PDF)
        material = limpiar_texto(item.get("material") or "")

        lineas.append({
            "id":               limpiar_texto(item.get("id", "")),
            "reference":        limpiar_texto(item.get("reference", "")),
            "name":             limpiar_texto(item.get("name", "")),
            "quantity":         convertir_a_float(item.get("quantity", "")),
            "total_linea":      convertir_a_float(item.get("total", "")),
            "observation":      limpiar_texto(item.get("observation", "")),
            "opening":          limpiar_texto(item.get("opening", "")),
            "size_x":           size_x,
            "size_y":           size_y,
            "size_z":           size_z,
            "price_total":      price_total,
            "model_door":       model_door,
            "material_door":    material_door,
            "material_cabinet": material_cabinet,
            "material":         material,
        })
    return resumen, lineas


# =========================================================
# PARSEAR PDF
# =========================================================

def extraer_texto_pdf(pdf_bytes):
    """Extrae texto probando varios modos de PyMuPDF para maximizar robustez."""
    pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    # Usamos "text" que devuelve el texto en orden de lectura con saltos de línea.
    texto = ""
    for page in pdf_doc:
        texto += page.get_text("text") + "\n"
    pdf_doc.close()
    return texto


def limpiar_lineas(texto):
    return [line.strip() for line in texto.splitlines() if line.strip()]


def extraer_pedido_pdf(texto):
    matches = re.findall(r'\b20\d{12}\b', texto)
    if matches:
        return matches[0]
    return ""


def parsear_cabecera_pdf(texto):
    lineas  = limpiar_lineas(texto)
    pedido  = extraer_pedido_pdf(texto)
    cliente = ""
    tienda  = ""
    try:
        idx    = lineas.index("Cliente:")
        bloque = lineas[idx + 1: idx + 6]
        if len(bloque) >= 5:
            cliente = limpiar_texto(bloque[3])
            tienda  = limpiar_texto(bloque[4])
    except ValueError:
        pass
    return {"pedido": pedido, "cliente": cliente, "tienda": tienda}


def extraer_importes_pdf(texto):
    lineas  = limpiar_lineas(texto)
    numeros = [l for l in lineas if re.fullmatch(r"\d+\.\d{2}", l)]
    if len(numeros) >= 3:
        return {
            "importe": convertir_a_float(numeros[-3]),
            "iva":     convertir_a_float(numeros[-2]),
            "total":   convertir_a_float(numeros[-1]),
        }
    return {"importe": None, "iva": None, "total": None}


# ---------- PARSER PDF ROBUSTO (3 estrategias) ----------

RE_POS_UNA_LINEA = re.compile(r"^(\d{1,3})\s+(\d{1,4})\s+(.+)$")

SECCIONES_VALIDAS = {"MUEBLES MURALES", "MUEBLES BAJOS", "MUEBLES ALTOS",
                      "REGLETAS", "COSTADOS", "COMPLEMENTOS",
                      "ACCESORIOS", "DECORATIVOS", "ELECTRODOMESTICOS", "ENCIMERAS"}

EXCLUIDOS_REF = {"POS", "MUEBLE", "UD.", "DESCRIPCION", "IMPORTE",
                 "MUEBLES", "BAJOS", "MURALES", "ALTOS",
                 "REGLETAS", "COSTADOS", "DECORATIVOS",
                 "COMPLEMENTOS", "ACCESORIOS", "ENCIMERAS",
                 "ELECTRODOMESTICOS", "IMPORTE", "I.V.A.",
                 "PRECIO", "TOTAL", "MODELO", "ACABADO",
                 "ARMAZÓN", "ARMAZON", "ZÓCALO", "ZOCALO",
                 "TIRADOR", "CAJÓN", "CAJON", "CLIENTE", "CREADO",
                 "REFERENCIA"}


def _parsear_estrategia_A_pos_una_linea(lineas):
    """Estrategia A: la POS viene como '1 1 ME2P40CX' en una sola línea."""
    indices_pos = []
    for idx, linea in enumerate(lineas):
        if linea.upper().startswith("POS MUEBLE"):
            continue
        m = RE_POS_UNA_LINEA.match(linea)
        if not m:
            continue
        pos, qty, ref = m.group(1), m.group(2), m.group(3).strip()
        if re.fullmatch(r"\d+\.\d{2}", ref):
            continue
        if ref.upper() in EXCLUIDOS_REF:
            continue
        indices_pos.append((idx, pos, qty, ref))
    return indices_pos


def _parsear_estrategia_B_pos_separada(lineas):
    """
    Estrategia B: POS en varias líneas consecutivas.
    Ejemplo:
        1
        1
        ME2P40CX
    """
    indices_pos = []
    i = 0
    while i < len(lineas) - 2:
        l1, l2, l3 = lineas[i], lineas[i+1], lineas[i+2]
        # l1 = número de posición (1-3 dígitos)
        # l2 = cantidad (1-4 dígitos)
        # l3 = referencia (no es número, no es sección, no es cabecera)
        if (re.fullmatch(r"\d{1,3}", l1)
                and re.fullmatch(r"\d{1,4}", l2)
                and not re.fullmatch(r"\d+(\.\d+)?", l3)
                and l3.upper() not in EXCLUIDOS_REF
                and l3.upper() not in SECCIONES_VALIDAS
                and not l3.upper().startswith("POS MUEBLE")
                and len(l3) >= 2):
            indices_pos.append((i, l1, l2, l3))
            i += 3
            continue
        i += 1
    return indices_pos


def _parsear_estrategia_C_ref_por_patron(lineas):
    """
    Estrategia C (último recurso): busca cualquier línea que PAREZCA una referencia
    (mayúsculas y números como ME2P40CX, RM230, CSM2, ZAL10.200, Puerta, Complemento)
    y que tenga cerca un 'L: x F: y A: z' e importe.
    """
    indices_pos = []
    for idx, linea in enumerate(lineas):
        if linea.upper() in EXCLUIDOS_REF:
            continue
        if linea.upper() in SECCIONES_VALIDAS:
            continue
        if re.fullmatch(r"\d+(\.\d+)?", linea):
            continue
        if linea.upper().startswith("POS MUEBLE"):
            continue
        # Referencias conocidas del ERP: códigos o palabras comunes
        # Reglas: al menos 3 caracteres, no todo símbolos
        if len(linea) < 3:
            continue
        # Descartar observaciones largas y descripciones
        if len(linea.split()) > 3:
            continue
        # Comprobar que en las 15 líneas siguientes hay un patrón L:.. F:.. A:..
        siguiente_15 = " ".join(lineas[idx+1: idx+16])
        if not re.search(r"L\s*:\s*\d+(?:\.\d+)?.*F\s*:\s*\d+(?:\.\d+)?.*A\s*:\s*\d+(?:\.\d+)?", siguiente_15, re.IGNORECASE):
            continue
        indices_pos.append((idx, "?", "?", linea))
    return indices_pos


def extraer_datos_bloque_pdf(bloque_lineas):
    """
    Extrae tamaño, opening y materiales del bloque de líneas que corresponden a UNA POS.
    Cada línea se examina por separado para que no se mezcle texto entre campos ni con otras POS.
    """
    size_x = size_y = size_z = None
    opening = ""
    model_door = ""
    material_door = ""
    material_cabinet = ""
    material = ""

    for linea in bloque_lineas:
        # Tamaño: "L: 400 F: 600 A: 2250"
        if size_x is None:
            m = re.search(r"L\s*:\s*(\d+(?:\.\d+)?)\s*F\s*:\s*(\d+(?:\.\d+)?)\s*A\s*:\s*(\d+(?:\.\d+)?)", linea, flags=re.IGNORECASE)
            if m:
                size_x = convertir_a_num(m.group(1))
                size_y = convertir_a_num(m.group(2))
                size_z = convertir_a_num(m.group(3))
                continue

        # Apertura: "M: Izquierda" / "M: Derecha"
        if not opening:
            m = re.search(r"^M\s*:\s*(Izquierda|Derecha)\s*$", linea, flags=re.IGNORECASE)
            if m:
                opening = m.group(1).capitalize()
                continue

        # "- Puertas: ESTRATO-196-FENIX VERDE KITAMI / "
        if not material_door and not model_door:
            m = re.search(r"^-?\s*Puertas\s*:\s*(.+?)\s*/?\s*$", linea, flags=re.IGNORECASE)
            if m:
                valor = m.group(1).strip()
                if "-" in valor:
                    partes = valor.split("-", 1)
                    model_door    = partes[0].strip()
                    material_door = partes[1].strip()
                else:
                    material_door = valor
                continue

        # "- Frente: 196-FENIX VERDE KITAMI"
        if not material_door:
            m = re.search(r"^-?\s*Frente\s*:\s*(.+)$", linea, flags=re.IGNORECASE)
            if m:
                material_door = m.group(1).strip()
                continue

        # "- Armazón:172-ROBLE" o "- Armazon:172-ROBLE"
        if not material_cabinet:
            m = re.search(r"^-?\s*Armaz[oó]n\s*:\s*(.+)$", linea, flags=re.IGNORECASE)
            if m:
                material_cabinet = m.group(1).strip()
                continue

        # "- Acabado:172-ROBLE" o "- Acabado:Laton"
        if not material:
            m = re.search(r"^-?\s*Acabado\s*:\s*(.+)$", linea, flags=re.IGNORECASE)
            if m:
                material = m.group(1).strip()
                continue

    return {
        "size_x": size_x, "size_y": size_y, "size_z": size_z,
        "opening": opening,
        "model_door": model_door,
        "material_door": material_door,
        "material_cabinet": material_cabinet,
        "material": material,
    }


def parsear_lineas_pdf(texto, debug_log=None):
    """
    Parser robusto que prueba 3 estrategias en orden.
    Si debug_log es una lista, registra información para mostrar al usuario.
    """
    lineas = limpiar_lineas(texto)

    if debug_log is not None:
        debug_log.append(f"Total líneas tras limpieza: {len(lineas)}")

    # Probar estrategia A
    indices = _parsear_estrategia_A_pos_una_linea(lineas)
    estrategia_usada = "A (POS en una línea)"

    if not indices:
        # Probar estrategia B
        indices = _parsear_estrategia_B_pos_separada(lineas)
        estrategia_usada = "B (POS en líneas separadas)"

    if not indices:
        # Probar estrategia C
        indices = _parsear_estrategia_C_ref_por_patron(lineas)
        estrategia_usada = "C (ref por patrón L:F:A)"

    if debug_log is not None:
        debug_log.append(f"Estrategia usada: {estrategia_usada}")
        debug_log.append(f"POS detectadas: {len(indices)}")

    if not indices:
        return []

    # Normalizar: indices es lista de (idx_linea, pos, qty, ref)
    resultados = []
    for n, (idx, pos, qty, ref) in enumerate(indices):
        fin = indices[n + 1][0] if n + 1 < len(indices) else len(lineas)
        # Empieza un poco después para coger descripción, tamaño, observaciones
        inicio = idx + 1
        bloque_lineas = lineas[inicio: fin]
        bloque_texto  = " ".join(bloque_lineas)

        descripcion = ""
        for l in bloque_lineas:
            if re.match(r"^L\s*:", l, flags=re.IGNORECASE): continue
            if re.match(r"^M\s*:", l, flags=re.IGNORECASE): continue
            if l.startswith("-"): continue
            if l.lower().startswith("observaciones"): continue
            if l.lower().startswith("variantes"): continue
            if re.fullmatch(r"\d+\.\d{2}", l): continue
            descripcion = l
            break

        importe = None
        for l in bloque_lineas:
            if re.fullmatch(r"\d+\.\d{2}", l):
                importe = convertir_a_float(l)
                break

        datos = extraer_datos_bloque_pdf(bloque_lineas)

        resultados.append({
            "pos":              str(pos),
            "reference":        limpiar_texto(ref),
            "description":      descripcion,
            "quantity":         convertir_a_float(qty) if qty != "?" else None,
            "importe_linea":    importe,
            "size_x":           datos["size_x"],
            "size_y":           datos["size_y"],
            "size_z":           datos["size_z"],
            "opening":          datos["opening"],
            "model_door":       datos["model_door"],
            "material_door":    datos["material_door"],
            "material_cabinet": datos["material_cabinet"],
            "material":         datos["material"],
        })
    return resultados


def indexar_por_referencia(lineas):
    refs = {}
    for item in lineas:
        ref = limpiar_upper(item.get("reference", ""))
        if ref and ref not in refs:
            refs[ref] = item
    return refs


def emparejar_lineas_por_id(json_lineas, pdf_lineas):
    pdf_disponibles = list(pdf_lineas)
    emparejados = []

    def quitar(pdf_item):
        for idx, p in enumerate(pdf_disponibles):
            if p is pdf_item:
                pdf_disponibles.pop(idx)
                return

    for j in json_lineas:
        ref_j = limpiar_upper(j.get("reference"))
        opn_j = limpiar_upper(j.get("opening"))
        sx, sy, sz = j.get("size_x"), j.get("size_y"), j.get("size_z")

        candidato = None

        if opn_j:
            for p in pdf_disponibles:
                if (limpiar_upper(p.get("reference")) == ref_j
                        and limpiar_upper(p.get("opening")) == opn_j):
                    candidato = p
                    break

        if candidato is None:
            for p in pdf_disponibles:
                if (limpiar_upper(p.get("reference")) == ref_j
                        and p.get("size_x") == sx
                        and p.get("size_y") == sy
                        and p.get("size_z") == sz):
                    candidato = p
                    break

        if candidato is None:
            for p in pdf_disponibles:
                if limpiar_upper(p.get("reference")) == ref_j:
                    candidato = p
                    break

        emparejados.append((j, candidato))
        if candidato is not None:
            quitar(candidato)

    return emparejados, pdf_disponibles


def formatear_tamano(x, y, z):
    def f(v):
        return str(v) if v is not None else "?"
    return f"{f(x)}×{f(y)}×{f(z)}"


# =========================================================
# PARSEAR EXCEL  (modo Excel ↔ PDF)
# =========================================================

def extraer_pedido_del_nombre(nombre_archivo):
    """Busca un número de pedido de 14 dígitos que empieza por 20 en el nombre del archivo."""
    m = re.search(r"\b(20\d{12})\b", nombre_archivo or "")
    return m.group(1) if m else ""


def parsear_excel(bytes_excel, nombre_archivo):
    """
    Lee el Excel que entrega el sistema. Estructura observada:
      Col 1: número de orden (1, 2, 3...)
      Col 2: referencia (ej. ME2P40CX, RM230, CSM2)
      Col 3: descripción
      Col 4: tamaño x  (largo, mm)
      Col 5: tamaño y  (fondo, mm)
      Col 6: tamaño z  (alto, mm)
      Col 7: precio UNITARIO
    No tiene cabecera de tabla ni datos de cliente / pedido.
    El número de pedido se saca del nombre del archivo.
    """
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(bytes_excel), data_only=True)
    # Tomamos la primera hoja (es lo único que trae el Excel observado)
    ws = wb[wb.sheetnames[0]]

    lineas = []
    contador_id = 0
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        ref = vals[1]
        if ref is None or not str(ref).strip():
            continue
        contador_id += 1
        lineas.append({
            "id":          f"xls_{contador_id}",
            "reference":   limpiar_texto(ref),
            "name":        limpiar_texto(vals[2]) if vals[2] is not None else "",
            "quantity":    None,            # el Excel no trae cantidad
            "total_linea": None,            # el Excel no trae importe total
            "observation": "",
            "opening":     "",              # tampoco trae opening
            "size_x":      convertir_a_num(vals[3]),
            "size_y":      convertir_a_num(vals[4]),
            "size_z":      convertir_a_num(vals[5]),
            "price_total": None,            # se rellena más abajo a partir de unidades * unitario
            "price_unit":  convertir_a_float(vals[6]),  # precio unitario que sí trae el Excel
        })

    resumen = {
        "pedido":   extraer_pedido_del_nombre(nombre_archivo),
        "cliente":  "",       # el Excel no trae cliente
        "tienda":   "",
        "proyecto": "",
        "importe":  None,
        "iva":      None,
        "total":    None,
    }
    return resumen, lineas


# =========================================================
# COMPARAR EXCEL ↔ PDF (modo específico)
# =========================================================

def comparar_excel_pdf(xls_resumen, xls_lineas, pdf_resumen, pdf_lineas):
    """
    Comparación específica Excel ↔ PDF.

    Lo que SÍ compara:
      - Que la lista de referencias del Excel y del PDF coincidan.
      - Que para cada línea emparejada el tamaño (x↔L, y↔F, z↔A) coincida.
      - Que el precio UNITARIO del Excel = importe_PDF / unidades_PDF.

    Lo que NO compara (porque el Excel no lo trae):
      - Cantidad (no está en el Excel)
      - Cabecera (importe, IVA, total, cliente, tienda)
      - Materiales / acabados
    """
    diferencias = []
    criticas    = []
    avisos      = []

    # Verificamos solo que el pedido del nombre del Excel = pedido del PDF
    # (esto en la práctica ya está garantizado por el emparejamiento, pero lo dejamos)
    if (xls_resumen.get("pedido") and pdf_resumen.get("pedido")
            and xls_resumen["pedido"] != pdf_resumen["pedido"]):
        criticas.append({
            "Campo": "Pedido", "JSON": xls_resumen["pedido"], "PDF": pdf_resumen["pedido"],
            "Diferencia": "",
            "Qué corregir": "El número de pedido del Excel no coincide con el del PDF.",
        })
        diferencias.append({
            "Gravedad": "🔴 Crítico", "Tipo": "Cabecera", "Campo": "Pedido",
            "Referencia": "CABECERA",
            "Valor JSON": xls_resumen["pedido"], "Valor PDF": pdf_resumen["pedido"],
            "Diferencia": "",
            "Qué corregir": "Pedidos diferentes en Excel y PDF.",
        })

    # Comparación por id (reutilizamos el emparejador del modo JSON)
    filas_tabla, criticas_id, avisos_id, diferencias_id = _comparar_por_id_xls(xls_lineas, pdf_lineas)
    criticas.extend(criticas_id)
    avisos.extend(avisos_id)
    diferencias.extend(diferencias_id)

    # Avisar de POS del PDF que se hayan quedado sin pareja en el Excel
    emparejados, huerfanos_pdf = emparejar_lineas_por_id(xls_lineas, pdf_lineas)
    for p in huerfanos_pdf:
        ref = p.get("reference") or ""
        avisos.append({
            "Campo": f"Extra en PDF — POS {p.get('pos')} ({ref})",
            "JSON": "—", "PDF": ref, "Diferencia": "",
            "Qué corregir": f"La línea POS {p.get('pos')} ({ref}) del PDF no aparece en el Excel.",
        })
        diferencias.append({
            "Gravedad": "🟡 Aviso", "Tipo": "Línea",
            "Campo": "Solo en PDF", "Referencia": ref,
            "Valor JSON": "", "Valor PDF": f"POS {p.get('pos')} — {ref}",
            "Diferencia": "",
            "Qué corregir": f"POS {p.get('pos')} ({ref}) está en PDF pero no en Excel.",
        })

    return diferencias, criticas, avisos, filas_tabla


def _comparar_por_id_xls(xls_lineas, pdf_lineas):
    """
    Versión específica de comparar_por_id para el modo Excel↔PDF.
    Compara solo lo que el Excel sí trae: tamaño y precio unitario.
    """
    filas_tabla = []
    criticas    = []
    avisos      = []
    diferencias = []

    emparejados, _ = emparejar_lineas_por_id(xls_lineas, pdf_lineas)

    for j, p in emparejados:
        id_xls   = j.get("id") or "—"
        ref      = j.get("reference") or "—"
        nombre   = j.get("name") or "—"
        tam_xls  = formatear_tamano(j.get("size_x"), j.get("size_y"), j.get("size_z"))
        precio_u_xls = j.get("price_unit")
        precio_u_xls_str = a_euro(precio_u_xls) if precio_u_xls is not None else "—"

        precio_elevado = precio_u_xls is not None and precio_u_xls > UMBRAL_PRECIO_LINEA

        if p is None:
            filas_tabla.append({
                "Ref":              ref,
                "Nombre":           nombre,
                "Tamaño Excel":     tam_xls,
                "Tamaño PDF":       "—",
                "Precio U. Excel":  precio_u_xls_str + (" ⚠️" if precio_elevado else ""),
                "Ud. PDF":          "—",
                "Importe PDF":      "—",
                "Precio U. PDF":    "—",
                "Dif. precio":      "—",
                "Estado":           "🔴 Sin pareja en PDF",
            })
            criticas.append({
                "Campo": f"Ref {ref}",
                "JSON": f"{ref} {tam_xls}",
                "PDF": "—",
                "Diferencia": "",
                "Qué corregir": f"La referencia {ref} ({tam_xls}) está en el Excel pero no se ha encontrado en el PDF.",
            })
            diferencias.append({
                "Gravedad": "🔴 Crítico", "Tipo": "Línea",
                "Campo": "Sin pareja PDF", "Referencia": ref,
                "Valor JSON": tam_xls, "Valor PDF": "",
                "Diferencia": "",
                "Qué corregir": f"{ref} en Excel no tiene pareja en el PDF.",
            })
            continue

        tam_pdf  = formatear_tamano(p.get("size_x"), p.get("size_y"), p.get("size_z"))
        ud_pdf   = p.get("quantity")
        imp_pdf  = convertir_a_float(p.get("importe_linea"))

        # Precio unitario del PDF = importe / unidades
        if imp_pdf is not None and ud_pdf is not None and ud_pdf != 0:
            precio_u_pdf = round(imp_pdf / ud_pdf, 2)
        else:
            precio_u_pdf = None
        precio_u_pdf_str = a_euro(precio_u_pdf) if precio_u_pdf is not None else "—"

        # Diferencia (PDF - Excel)
        if precio_u_xls is not None and precio_u_pdf is not None:
            dif = round(precio_u_pdf - precio_u_xls, 2)
            dif_str = a_euro(dif)
        else:
            dif = None
            dif_str = "—"

        # Tamaño: comparar con tolerancia mínima por decimales
        diffs_ejes = []
        for eje, vj, vp in [("x/L", j.get("size_x"), p.get("size_x")),
                             ("y/F", j.get("size_y"), p.get("size_y")),
                             ("z/A", j.get("size_z"), p.get("size_z"))]:
            if vj is None or vp is None:
                if vj != vp:
                    diffs_ejes.append(f"{eje}: Excel={vj} / PDF={vp}")
            else:
                # Tolerancia de 0.5 mm por si hay redondeos en el PDF
                if abs(float(vj) - float(vp)) > 0.5:
                    diffs_ejes.append(f"{eje}: Excel={vj} / PDF={vp}")

        # Estado
        motivos = []
        if diffs_ejes:                            motivos.append("tamaño")
        if dif is not None and abs(dif) > 0.01:   motivos.append("precio")

        estado = "🔴 " + " + ".join(motivos).capitalize() + " distinto" if motivos else "🟢 OK"

        # Registrar diferencias
        if diffs_ejes:
            detalle = " · ".join(diffs_ejes)
            criticas.append({
                "Campo": f"Tamaño — {ref}",
                "JSON": tam_xls, "PDF": tam_pdf,
                "Diferencia": detalle,
                "Qué corregir": f"Tamaño distinto en {ref}: {detalle}",
            })
            diferencias.append({
                "Gravedad": "🔴 Crítico", "Tipo": "Línea",
                "Campo": "Tamaño", "Referencia": ref,
                "Valor JSON": tam_xls, "Valor PDF": tam_pdf,
                "Diferencia": detalle,
                "Qué corregir": f"Tamaño distinto en {ref}: {detalle}",
            })

        if dif is not None and abs(dif) > 0.01:
            criticas.append({
                "Campo": f"Precio unitario — {ref}",
                "JSON": precio_u_xls_str, "PDF": precio_u_pdf_str,
                "Diferencia": dif_str,
                "Qué corregir": f"Precio unitario distinto en {ref}: Excel={precio_u_xls_str} / PDF={precio_u_pdf_str} (Dif: {dif_str})",
            })
            diferencias.append({
                "Gravedad": "🔴 Crítico", "Tipo": "Línea",
                "Campo": "Precio unitario", "Referencia": ref,
                "Valor JSON": precio_u_xls_str, "Valor PDF": precio_u_pdf_str,
                "Diferencia": dif_str,
                "Qué corregir": f"Precio unitario distinto en {ref}.",
            })

        ud_pdf_str = str(int(ud_pdf)) if ud_pdf is not None and ud_pdf == int(ud_pdf) else (str(ud_pdf) if ud_pdf is not None else "—")
        imp_pdf_str = a_euro(imp_pdf) if imp_pdf is not None else "—"

        filas_tabla.append({
            "Ref":              ref,
            "Nombre":           nombre,
            "Tamaño Excel":     tam_xls,
            "Tamaño PDF":       tam_pdf,
            "Precio U. Excel":  precio_u_xls_str + (" ⚠️" if precio_elevado else ""),
            "Ud. PDF":          ud_pdf_str,
            "Importe PDF":      imp_pdf_str,
            "Precio U. PDF":    precio_u_pdf_str,
            "Dif. precio":      dif_str,
            "Estado":           estado,
        })

        if precio_elevado:
            avisos.append({
                "Campo": f"Precio elevado — {ref}",
                "JSON": precio_u_xls_str, "PDF": "",
                "Diferencia": "",
                "Qué corregir": f"El precio unitario de {ref} es {precio_u_xls_str} (> {UMBRAL_PRECIO_LINEA:.0f} €). Revisar.",
            })
            diferencias.append({
                "Gravedad": "🟡 Aviso", "Tipo": "Línea",
                "Campo": "Precio elevado (>4000 €)", "Referencia": ref,
                "Valor JSON": precio_u_xls_str, "Valor PDF": "",
                "Diferencia": "",
                "Qué corregir": f"Precio unitario {precio_u_xls_str} supera el umbral.",
            })

    return filas_tabla, criticas, avisos, diferencias


# =========================================================
# COMPARAR UN PAR
# =========================================================

def comparar_par(json_resumen, json_lineas, pdf_resumen, pdf_lineas):
    diferencias = []
    criticas    = []
    avisos      = []

    campos = [
        ("Cliente",  json_resumen["cliente"],  pdf_resumen["cliente"],  False),
        ("Pedido",   json_resumen["pedido"],   pdf_resumen["pedido"],   False),
        ("Tienda",   json_resumen["tienda"],   pdf_resumen["tienda"],   False),
        ("Importe",  json_resumen["importe"],  pdf_resumen["importe"],  True),
        ("IVA",      json_resumen["iva"],      pdf_resumen["iva"],      True),
        ("Total",    json_resumen["total"],    pdf_resumen["total"],    True),
    ]

    for campo, vj, vp, es_num in campos:
        hay_diff = son_numeros_distintos(vj, vp) if es_num else son_textos_distintos(vj, vp)
        if hay_diff:
            diff_str = ""
            if es_num:
                fj = convertir_a_float(vj)
                fp = convertir_a_float(vp)
                if fj is not None and fp is not None:
                    diff_str = a_euro(round(fp - fj, 2))
                criticas.append({"Campo": campo, "JSON": a_euro(vj), "PDF": a_euro(vp), "Diferencia": diff_str, "Qué corregir": f"El {campo} no coincide. Revisar."})
                gravedad = "🔴 Crítico"
            else:
                avisos.append({"Campo": campo, "JSON": vj, "PDF": vp, "Diferencia": "", "Qué corregir": f"El {campo} no coincide. Verificar."})
                gravedad = "🟡 Aviso"
            diferencias.append({"Gravedad": gravedad, "Tipo": "Cabecera", "Campo": campo, "Referencia": "CABECERA",
                                 "Valor JSON": a_euro(vj) if es_num else vj, "Valor PDF": a_euro(vp) if es_num else vp,
                                 "Diferencia": diff_str, "Qué corregir": f"{campo} no coincide."})

    refs_json = indexar_por_referencia(json_lineas)
    refs_pdf  = indexar_por_referencia(pdf_lineas)
    solo_json = sorted(set(refs_json.keys()) - set(refs_pdf.keys()))
    solo_pdf  = sorted(set(refs_pdf.keys())  - set(refs_json.keys()))

    # Avisar si hay referencias que están en PDF pero no en el JSON (líneas sobrantes).
    # Las referencias que están en JSON pero no en PDF se detectan luego por id
    # (saldrán como "Sin pareja en PDF" en la tabla).
    for ref in solo_pdf:
        avisos.append({"Campo": "Extra en PDF", "JSON": "—", "PDF": ref, "Diferencia": "", "Qué corregir": f"Referencia {ref} aparece en el PDF pero no en el JSON."})
        diferencias.append({"Gravedad": "🟡 Aviso", "Tipo": "Línea", "Campo": "Solo en PDF", "Referencia": ref,
                             "Valor JSON": "", "Valor PDF": ref, "Diferencia": "", "Qué corregir": f"Referencia {ref} aparece solo en el PDF."})

    comparacion_id, criticas_id, avisos_id, diferencias_id = comparar_por_id(json_lineas, pdf_lineas)
    criticas.extend(criticas_id)
    avisos.extend(avisos_id)
    diferencias.extend(diferencias_id)

    return diferencias, criticas, avisos, comparacion_id


def comparar_por_id(json_lineas, pdf_lineas):
    filas_tabla = []
    criticas    = []
    avisos      = []
    diferencias = []

    emparejados, _pdf_huerfanos = emparejar_lineas_por_id(json_lineas, pdf_lineas)

    for j, p in emparejados:
        id_json  = j.get("id") or "—"
        ref      = j.get("reference") or "—"
        nombre   = j.get("name") or "—"
        opening  = j.get("opening") or ""
        tam_json = formatear_tamano(j.get("size_x"), j.get("size_y"), j.get("size_z"))
        precio_j = j.get("price_total")
        precio_j_str = a_euro(precio_j) if precio_j is not None else "—"
        cant_j   = j.get("quantity")
        cant_j_str = str(int(cant_j)) if cant_j is not None and cant_j == int(cant_j) else (str(cant_j) if cant_j is not None else "—")

        precio_elevado = precio_j is not None and precio_j > UMBRAL_PRECIO_LINEA

        if p is None:
            filas_tabla.append({
                "ID":           id_json,
                "Referencia":   ref,
                "Nombre":       nombre + (f" ({opening})" if opening else ""),
                "Cant. JSON":   cant_j_str,
                "Cant. PDF":    "—",
                "Tamaño JSON":  tam_json,
                "Tamaño PDF":   "—",
                "Precio JSON":  precio_j_str + (" ⚠️" if precio_elevado else ""),
                "Precio PDF":   "—",
                "Dif. precio":  "—",
                "Estado":       "🔴 Sin pareja en PDF",
            })
            criticas.append({
                "Campo": f"ID {id_json} ({ref})",
                "JSON": f"{ref} {tam_json}",
                "PDF": "—",
                "Diferencia": "",
                "Qué corregir": f"El id {id_json} ({ref}) no se ha emparejado con ninguna línea del PDF.",
            })
            diferencias.append({
                "Gravedad": "🔴 Crítico", "Tipo": "ID",
                "Campo": "Sin pareja PDF", "Referencia": f"{id_json} / {ref}",
                "Valor JSON": tam_json, "Valor PDF": "",
                "Diferencia": "",
                "Qué corregir": f"ID {id_json} ({ref}) no encontrado en el PDF.",
            })
        else:
            tam_pdf = formatear_tamano(p.get("size_x"), p.get("size_y"), p.get("size_z"))
            precio_p = convertir_a_float(p.get("importe_linea"))
            precio_p_str = a_euro(precio_p) if precio_p is not None else "—"
            cant_p   = p.get("quantity")
            cant_p_str = str(int(cant_p)) if cant_p is not None and cant_p == int(cant_p) else (str(cant_p) if cant_p is not None else "—")

            if precio_j is not None and precio_p is not None:
                dif_precio = round(precio_p - precio_j, 2)
                dif_precio_str = a_euro(dif_precio)
            else:
                dif_precio = None
                dif_precio_str = "—"

            # --- Comparación de tamaño (eje a eje) ---
            diffs_ejes = []
            for eje, vj, vp in [
                ("x/L", j.get("size_x"), p.get("size_x")),
                ("y/F", j.get("size_y"), p.get("size_y")),
                ("z/A", j.get("size_z"), p.get("size_z")),
            ]:
                if vj != vp:
                    diffs_ejes.append(f"{eje}: JSON={vj} / PDF={vp}")

            # --- Comparación de cantidad ---
            cant_distinta = son_numeros_distintos(cant_j, cant_p)

            # --- Estado resumen ---
            motivos = []
            if diffs_ejes:      motivos.append("tamaño")
            if cant_distinta:   motivos.append("cantidad")
            if dif_precio is not None and dif_precio != 0: motivos.append("precio")

            if motivos:
                estado = "🔴 " + " + ".join(motivos).capitalize() + " distinto"
            else:
                estado = "🟢 OK"

            # --- Registrar cada diferencia individualmente ---
            if diffs_ejes:
                detalle = " · ".join(diffs_ejes)
                criticas.append({
                    "Campo": f"Tamaño — {id_json} ({ref})",
                    "JSON": tam_json, "PDF": tam_pdf,
                    "Diferencia": detalle,
                    "Qué corregir": f"Tamaño distinto en {ref} (id {id_json}): {detalle}",
                })
                diferencias.append({
                    "Gravedad": "🔴 Crítico", "Tipo": "ID",
                    "Campo": "Tamaño", "Referencia": f"{id_json} / {ref}",
                    "Valor JSON": tam_json, "Valor PDF": tam_pdf,
                    "Diferencia": detalle,
                    "Qué corregir": f"Tamaño distinto en {ref}: {detalle}",
                })

            if cant_distinta:
                criticas.append({
                    "Campo": f"Cantidad — {id_json} ({ref})",
                    "JSON": cant_j_str, "PDF": cant_p_str,
                    "Diferencia": "",
                    "Qué corregir": f"Cantidad distinta en {ref} (id {id_json}): JSON={cant_j_str} / PDF={cant_p_str}",
                })
                diferencias.append({
                    "Gravedad": "🔴 Crítico", "Tipo": "ID",
                    "Campo": "Cantidad", "Referencia": f"{id_json} / {ref}",
                    "Valor JSON": cant_j_str, "Valor PDF": cant_p_str,
                    "Diferencia": "",
                    "Qué corregir": f"Cantidad distinta en {ref}.",
                })

            if dif_precio is not None and dif_precio != 0:
                criticas.append({
                    "Campo": f"Precio línea — {id_json} ({ref})",
                    "JSON": precio_j_str, "PDF": precio_p_str,
                    "Diferencia": dif_precio_str,
                    "Qué corregir": f"Precio de línea {ref} (id {id_json}): JSON={precio_j_str} PDF={precio_p_str} Dif={dif_precio_str}",
                })
                diferencias.append({
                    "Gravedad": "🔴 Crítico", "Tipo": "ID",
                    "Campo": "Precio línea", "Referencia": f"{id_json} / {ref}",
                    "Valor JSON": precio_j_str, "Valor PDF": precio_p_str,
                    "Diferencia": dif_precio_str,
                    "Qué corregir": f"Precio de línea distinto en {ref}.",
                })

            filas_tabla.append({
                "ID":            id_json,
                "Referencia":    ref,
                "Nombre":        nombre + (f" ({opening})" if opening else ""),
                "Cant. JSON":    cant_j_str,
                "Cant. PDF":     cant_p_str,
                "Tamaño JSON":   tam_json,
                "Tamaño PDF":    tam_pdf,
                "Precio JSON":   precio_j_str + (" ⚠️" if precio_elevado else ""),
                "Precio PDF":    precio_p_str,
                "Dif. precio":   dif_precio_str,
                "Estado":        estado,
            })

        if precio_elevado:
            avisos.append({
                "Campo": f"Precio elevado — {id_json} ({ref})",
                "JSON": precio_j_str,
                "PDF": "",
                "Diferencia": "",
                "Qué corregir": f"El id {id_json} ({ref}) tiene un precio de {precio_j_str} (> {UMBRAL_PRECIO_LINEA:.0f} €). Revisar posible valor erróneo.",
            })
            diferencias.append({
                "Gravedad": "🟡 Aviso", "Tipo": "ID",
                "Campo": "Precio elevado (>4000 €)", "Referencia": f"{id_json} / {ref}",
                "Valor JSON": precio_j_str, "Valor PDF": "",
                "Diferencia": "",
                "Qué corregir": f"Precio de línea {precio_j_str} supera el umbral de {UMBRAL_PRECIO_LINEA:.0f} €. Revisar.",
            })

    return filas_tabla, criticas, avisos, diferencias


# =========================================================
# MOSTRAR RESULTADO DE UN CLIENTE
# =========================================================

def mostrar_resultado(pedido, cliente, json_resumen, pdf_resumen, diferencias, criticas, avisos, comparacion_id,
                      pdf_debug_texto=None, pdf_debug_log=None, pdf_lineas=None):
    n_crit  = len(criticas)
    n_avis  = len(avisos)
    n_total = n_crit + n_avis

    with st.expander(f"📦 Pedido {pedido} — {cliente}", expanded=(n_crit > 0)):
        if n_total == 0:
            st.markdown('<div class="semaforo-verde">✅ TODO CORRECTO</div>', unsafe_allow_html=True)
        elif n_crit > 0:
            st.markdown(f'<div class="semaforo-rojo">🔴 {n_crit} crítica(s) — {n_avis} aviso(s)</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="semaforo-amarillo">🟡 {n_avis} aviso(s) — Revisar</div>', unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        c1.metric("Cliente", json_resumen["cliente"] or pdf_resumen["cliente"] or "—")
        c2.metric("Pedido",  pedido)
        c3.metric("Tienda",  json_resumen["tienda"]  or pdf_resumen["tienda"]  or "—")

        c4, c5, c6 = st.columns(3)
        for col, campo, vj, vp in [(c4, "Importe", json_resumen["importe"], pdf_resumen["importe"]),
                                    (c5, "IVA",     json_resumen["iva"],     pdf_resumen["iva"]),
                                    (c6, "Total",   json_resumen["total"],   pdf_resumen["total"])]:
            fj = convertir_a_float(vj)
            fp = convertir_a_float(vp)
            delta = round(fp - fj, 2) if fj and fp else None
            col.metric(f"{campo} JSON", a_euro(fj), delta=f"{delta} €" if delta else None)

        if criticas:
            st.markdown("#### 🔴 Diferencias críticas")
            for d in criticas:
                st.error(f"**{d['Campo']}** → JSON: `{d['JSON']}` | PDF: `{d['PDF']}` {'| Dif: ' + d['Diferencia'] if d['Diferencia'] else ''}")
                st.caption(f"💡 {d['Qué corregir']}")

        if avisos:
            st.markdown("#### 🟡 Avisos")
            for d in avisos:
                st.warning(f"**{d['Campo']}** → JSON: `{d['JSON']}` | PDF: `{d['PDF']}`")
                st.caption(f"💡 {d['Qué corregir']}")

        if comparacion_id:
            st.markdown("#### 🔍 Comparación por ID (tamaño y precio por línea)")
            st.caption(
                "Cada fila corresponde a un `id` del JSON. Se compara el tamaño "
                "(x↔L, y↔F, z↔A) y el precio de línea con el PDF. "
                f"El icono ⚠️ junto al precio indica que supera {UMBRAL_PRECIO_LINEA:.0f} € "
                "(posible valor erróneo)."
            )
            df_id = pd.DataFrame(comparacion_id)
            st.dataframe(df_id, use_container_width=True, hide_index=True)

        # ---------- PANEL DE DEPURACIÓN ----------
        with st.expander("🛠️ Depuración del PDF (ábreme si algo no cuadra)"):
            st.write("**Log del parser:**")
            if pdf_debug_log:
                for msg in pdf_debug_log:
                    st.code(msg)
            st.write(f"**Líneas detectadas del PDF: {len(pdf_lineas) if pdf_lineas else 0}**")
            if pdf_lineas:
                st.dataframe(pd.DataFrame(pdf_lineas), use_container_width=True, hide_index=True)
            st.write("**Texto extraído (primeras 60 líneas):**")
            if pdf_debug_texto:
                lineas_preview = pdf_debug_texto.splitlines()[:60]
                st.code("\n".join(f"{i:3d}: {l}" for i, l in enumerate(lineas_preview) if l.strip()))

        if diferencias:
            st.markdown("#### 📋 Tabla completa de diferencias")
            st.dataframe(pd.DataFrame(diferencias), use_container_width=True, hide_index=True)
            excel = crear_excel_en_memoria(pd.DataFrame(diferencias))
            st.download_button(
                label=f"📥 Descargar Excel — Pedido {pedido}",
                data=excel,
                file_name=f"diferencias_pedido_{pedido}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"excel_{pedido}"
            )


# =========================================================
# INTERFAZ PRINCIPAL
# =========================================================

st.title("📋 Comparador de pedidos vs PDF")
st.write("Sube los archivos y la app los empareja automáticamente por número de pedido.")

# ---------- Selector de modo ----------
modo = st.radio(
    "**Tipo de comparación:**",
    options=["📂 JSON ↔ PDF", "📊 Excel ↔ PDF"],
    horizontal=True,
    help="JSON ↔ PDF: ficheros internos. Excel ↔ PDF: el Excel del sistema (debe llamarse con el número de pedido, ej. '20260430473073.xls')."
)

modo_excel = modo.startswith("📊")

st.markdown("---")

# ---------- Uploaders ----------
col1, col2 = st.columns(2)

if modo_excel:
    with col1:
        xls_files = st.file_uploader(
            "📊 Archivos Excel (el nombre debe ser el número de pedido)",
            type=["xls", "xlsx"],
            accept_multiple_files=True,
        )
    with col2:
        pdf_files = st.file_uploader(
            "📄 Archivos PDF",
            type=["pdf"],
            accept_multiple_files=True,
        )
    json_files = None
else:
    with col1:
        json_files = st.file_uploader(
            "📂 Archivos JSON",
            type=["json"],
            accept_multiple_files=True,
        )
    with col2:
        pdf_files = st.file_uploader(
            "📄 Archivos PDF",
            type=["pdf"],
            accept_multiple_files=True,
        )
    xls_files = None


# =====================================================
# MODO JSON ↔ PDF
# =====================================================
if not modo_excel and json_files and pdf_files:
    st.markdown("---")

    jsons = {}
    for f in json_files:
        try:
            data = json.load(f)
            resumen, lineas = parsear_json(data)
            if resumen["pedido"]:
                jsons[resumen["pedido"]] = (resumen, lineas)
            else:
                st.warning(f"⚠️ {f.name} no tiene número de pedido.")
        except Exception as e:
            st.error(f"Error en {f.name}: {e}")

    pdfs = {}
    for f in pdf_files:
        try:
            texto    = extraer_texto_pdf(f.read())
            cabecera = parsear_cabecera_pdf(texto)
            importes = extraer_importes_pdf(texto)
            debug_log = []
            lineas   = parsear_lineas_pdf(texto, debug_log=debug_log)
            pedido   = cabecera["pedido"]
            if pedido:
                pdfs[pedido] = (
                    {
                        "pedido":  pedido,
                        "cliente": cabecera["cliente"],
                        "tienda":  cabecera["tienda"],
                        "importe": importes["importe"],
                        "iva":     importes["iva"],
                        "total":   importes["total"],
                    },
                    lineas,
                    texto,
                    debug_log,
                )
            else:
                st.warning(f"⚠️ {f.name} no tiene número de pedido reconocible.")
        except Exception as e:
            st.error(f"Error en {f.name}: {e}")

    emparejados = sorted(set(jsons.keys()) & set(pdfs.keys()))
    sin_pdf     = sorted(set(jsons.keys()) - set(pdfs.keys()))
    sin_json    = sorted(set(pdfs.keys())  - set(jsons.keys()))

    st.markdown(f"### 📊 {len(emparejados)} pedido(s) comparado(s)")
    c1, c2, c3 = st.columns(3)
    c1.metric("Pares encontrados", len(emparejados))
    c2.metric("JSON sin PDF",      len(sin_pdf))
    c3.metric("PDF sin JSON",      len(sin_json))

    for p in sin_pdf:
        st.markdown(f'<div class="sin-pareja">⚠️ Pedido <b>{p}</b> — tiene JSON pero no se encontró su PDF</div>', unsafe_allow_html=True)
    for p in sin_json:
        st.markdown(f'<div class="sin-pareja">⚠️ Pedido <b>{p}</b> — tiene PDF pero no se encontró su JSON</div>', unsafe_allow_html=True)

    st.markdown("---")

    total_crit = 0
    total_avis = 0

    for pedido in emparejados:
        json_resumen, json_lineas = jsons[pedido]
        pdf_resumen, pdf_lineas, pdf_texto, pdf_log = pdfs[pedido]
        difs, criticas, avisos, comparacion_id = comparar_par(json_resumen, json_lineas, pdf_resumen, pdf_lineas)
        total_crit += len(criticas)
        total_avis += len(avisos)
        mostrar_resultado(
            pedido,
            json_resumen["cliente"] or pdf_resumen["cliente"],
            json_resumen, pdf_resumen,
            difs, criticas, avisos,
            comparacion_id,
            pdf_debug_texto=pdf_texto,
            pdf_debug_log=pdf_log,
            pdf_lineas=pdf_lineas,
        )

    st.markdown("---")
    st.markdown("### 🚦 Estado general")
    if total_crit > 0:
        st.markdown(f'<div class="semaforo-rojo">🔴 HAY PROBLEMAS — {total_crit} diferencia(s) crítica(s) en total</div>', unsafe_allow_html=True)
    elif total_avis > 0:
        st.markdown(f'<div class="semaforo-amarillo">🟡 REVISAR — {total_avis} aviso(s) en total</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="semaforo-verde">✅ TODOS LOS PEDIDOS CORRECTOS</div>', unsafe_allow_html=True)


# =====================================================
# MODO EXCEL ↔ PDF
# =====================================================
if modo_excel and xls_files and pdf_files:
    st.markdown("---")

    # Leer Excels
    xls_dict = {}
    for f in xls_files:
        try:
            contenido = f.read()
            resumen, lineas = parsear_excel(contenido, f.name)
            pedido = resumen["pedido"]
            if pedido:
                xls_dict[pedido] = (resumen, lineas, f.name)
            else:
                st.warning(f"⚠️ **{f.name}** — no se ha detectado un número de pedido en el nombre del archivo. "
                           "El nombre debe contener el número de 14 dígitos (ej: `20260430473073.xls`).")
        except Exception as e:
            st.error(f"Error leyendo Excel {f.name}: {e}")

    # Leer PDFs
    pdfs = {}
    for f in pdf_files:
        try:
            texto    = extraer_texto_pdf(f.read())
            cabecera = parsear_cabecera_pdf(texto)
            importes = extraer_importes_pdf(texto)
            debug_log = []
            lineas   = parsear_lineas_pdf(texto, debug_log=debug_log)
            pedido   = cabecera["pedido"]
            if pedido:
                pdfs[pedido] = (
                    {
                        "pedido":  pedido,
                        "cliente": cabecera["cliente"],
                        "tienda":  cabecera["tienda"],
                        "importe": importes["importe"],
                        "iva":     importes["iva"],
                        "total":   importes["total"],
                    },
                    lineas, texto, debug_log,
                )
            else:
                st.warning(f"⚠️ {f.name} no tiene número de pedido reconocible.")
        except Exception as e:
            st.error(f"Error en {f.name}: {e}")

    emparejados = sorted(set(xls_dict.keys()) & set(pdfs.keys()))
    sin_pdf     = sorted(set(xls_dict.keys()) - set(pdfs.keys()))
    sin_xls     = sorted(set(pdfs.keys())     - set(xls_dict.keys()))

    st.markdown(f"### 📊 {len(emparejados)} pedido(s) comparado(s)")
    c1, c2, c3 = st.columns(3)
    c1.metric("Pares encontrados", len(emparejados))
    c2.metric("Excel sin PDF",     len(sin_pdf))
    c3.metric("PDF sin Excel",     len(sin_xls))

    for p in sin_pdf:
        st.markdown(f'<div class="sin-pareja">⚠️ Pedido <b>{p}</b> — tiene Excel pero no se encontró su PDF</div>', unsafe_allow_html=True)
    for p in sin_xls:
        st.markdown(f'<div class="sin-pareja">⚠️ Pedido <b>{p}</b> — tiene PDF pero no se encontró su Excel</div>', unsafe_allow_html=True)

    st.markdown("---")

    total_crit = 0
    total_avis = 0

    for pedido in emparejados:
        xls_resumen, xls_lineas, xls_nombre = xls_dict[pedido]
        pdf_resumen, pdf_lineas, pdf_texto, pdf_log = pdfs[pedido]
        difs, criticas, avisos, comparacion = comparar_excel_pdf(xls_resumen, xls_lineas, pdf_resumen, pdf_lineas)
        total_crit += len(criticas)
        total_avis += len(avisos)
        cliente = pdf_resumen["cliente"] or "—"

        n_crit, n_avis = len(criticas), len(avisos)
        with st.expander(f"📦 Pedido {pedido} — {cliente}", expanded=(n_crit > 0)):
            if n_crit == 0 and n_avis == 0:
                st.markdown('<div class="semaforo-verde">✅ TODO CORRECTO</div>', unsafe_allow_html=True)
            elif n_crit > 0:
                st.markdown(f'<div class="semaforo-rojo">🔴 {n_crit} crítica(s) — {n_avis} aviso(s)</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="semaforo-amarillo">🟡 {n_avis} aviso(s) — Revisar</div>', unsafe_allow_html=True)

            c1, c2, c3 = st.columns(3)
            c1.metric("Cliente", cliente)
            c2.metric("Pedido",  pedido)
            c3.metric("Archivo Excel", xls_nombre)

            if criticas:
                st.markdown("#### 🔴 Diferencias críticas")
                for d in criticas:
                    extra = f" | Dif: {d['Diferencia']}" if d.get('Diferencia') else ""
                    st.error(f"**{d['Campo']}** → Excel: `{d['JSON']}` | PDF: `{d['PDF']}`{extra}")
                    st.caption(f"💡 {d['Qué corregir']}")

            if avisos:
                st.markdown("#### 🟡 Avisos")
                for d in avisos:
                    st.warning(f"**{d['Campo']}** → Excel: `{d['JSON']}` | PDF: `{d['PDF']}`")
                    st.caption(f"💡 {d['Qué corregir']}")

            if comparacion:
                st.markdown("#### 🔍 Comparación línea a línea")
                st.caption(
                    "El precio unitario del Excel se compara contra el del PDF (importe ÷ unidades). "
                    "El Excel no trae cantidad ni cabecera; eso no se compara."
                )
                st.dataframe(pd.DataFrame(comparacion), use_container_width=True, hide_index=True)

            with st.expander("🛠️ Depuración del PDF (ábreme si algo no cuadra)"):
                st.write("**Log del parser:**")
                for msg in pdf_log:
                    st.code(msg)
                st.write(f"**Líneas detectadas del PDF: {len(pdf_lineas)}**")
                if pdf_lineas:
                    st.dataframe(pd.DataFrame(pdf_lineas), use_container_width=True, hide_index=True)
                st.write("**Texto extraído (primeras 60 líneas):**")
                lineas_preview = pdf_texto.splitlines()[:60]
                st.code("\n".join(f"{i:3d}: {l}" for i, l in enumerate(lineas_preview) if l.strip()))

            if difs:
                st.markdown("#### 📋 Tabla completa de diferencias")
                st.dataframe(pd.DataFrame(difs), use_container_width=True, hide_index=True)
                excel_bytes = crear_excel_en_memoria(pd.DataFrame(difs))
                st.download_button(
                    label=f"📥 Descargar Excel — Pedido {pedido}",
                    data=excel_bytes,
                    file_name=f"diferencias_pedido_{pedido}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"excel_xls_{pedido}",
                )

    st.markdown("---")
    st.markdown("### 🚦 Estado general")
    if total_crit > 0:
        st.markdown(f'<div class="semaforo-rojo">🔴 HAY PROBLEMAS — {total_crit} diferencia(s) crítica(s) en total</div>', unsafe_allow_html=True)
    elif total_avis > 0:
        st.markdown(f'<div class="semaforo-amarillo">🟡 REVISAR — {total_avis} aviso(s) en total</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="semaforo-verde">✅ TODOS LOS PEDIDOS CORRECTOS</div>', unsafe_allow_html=True)
